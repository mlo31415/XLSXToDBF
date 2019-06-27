import dbf
import openpyxl
import os

from openpyxl import load_workbook
wb = load_workbook(filename = '..//..//books.DBF.xlsx')
ws = wb['books']
print(ws['C18'].value)

table = dbf.Table('books', 'NSCN N(5,0); AUTHOR C(40); TITLE C(100); LOGICAL C(1); PHYSICAL C(1); NUMAUTH C(1); OTHER C(1); CONDITION C(1); TYPE C(1); SERIES C(5)')
print('db definition created with field names:', table.field_names)

table.open(mode=dbf.READ_WRITE)

def ToInt(c):
    if c is None or c.value is None or (type(c.value) is str and len(c.value) == 0):
        return None
    if type(c.value) is str:
        return int(c.value)
    return c.value

def ToStr(c):
    if c is None or c.value is None or (type(c.value) is str and len(c.value) == 0):
        return None
    if type(c.value) is str:
        return c.value
    try:
        return str(c.value)
    except:
        return None


skip=True
for row in ws.rows:
    if skip:
        skip=False
        continue
    print(row[0].value)
    table.append((ToInt(row[0]), ToStr(row[1]), ToStr(row[2]), ToStr(row[3]), ToStr(row[4]), ToStr(row[5]), ToStr(row[6]), ToStr(row[7]), ToStr(row[8]), ToStr(row[9])))
    i=0


